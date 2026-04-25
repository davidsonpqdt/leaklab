"""Classifica arquivos XML iPoker em bounty_ko / mystery / vanilla.

Heuristica:
  - Nome do torneio contem 'MYSTERY'                                -> Mystery
  - Nome contem BOUNTY/KNOCKOUT/PROGRESSIVE ou token 'KO'           -> Bounty/KO
  - Buyin tem 3 partes (X + Y + rake)                               -> Bounty/KO (fallback)
  - Caso contrario                                                  -> Vanilla

Tambem produz xlsx de resumo agrupado por codigo de torneio.

Uso:
    python tournament_classifier.py /pasta/dos/xmls --output-dir /pasta/saida
"""
from __future__ import annotations
import os, re, glob, shutil, argparse
from collections import defaultdict
from xml.etree import ElementTree as ET


def classify(buyin: str, tname: str) -> str:
    name = (tname or "").upper()
    if "MYSTERY" in name:
        return "Mystery"
    for kw in ("BOUNTY", "KNOCKOUT", "PROGRESSIVE"):
        if kw in name:
            return "Bounty/KO"
    if re.search(r"\bKO\b", name):
        return "Bounty/KO"
    parts = [p.strip() for p in (buyin or "").split("+") if p.strip()]
    if len(parts) >= 3:
        return "Bounty/KO"
    return "Vanilla"


def parse_metadata(path: str) -> dict | None:
    """Le metadados do torneio do <general> raiz. Retorna None se XML invalido."""
    try:
        tree = ET.parse(path)
    except Exception:
        return None
    g = tree.getroot().find("general")
    if g is None:
        return None
    return {
        "file": os.path.basename(path),
        "path": path,
        "tcode": (g.findtext("tournamentcode") or "").strip(),
        "tname": (g.findtext("tournamentname") or "").strip(),
        "buyin": (g.findtext("buyin") or "").strip(),
        "totalbuyin": (g.findtext("totalbuyin") or "").strip(),
        "startdate": (g.findtext("startdate") or "").strip(),
        "tablesize": (g.findtext("tablesize") or "").strip(),
    }


def collect_metadata(folder: str) -> list[dict]:
    out = []
    for path in sorted(glob.glob(os.path.join(folder, "*.xml"))):
        m = parse_metadata(path)
        if m is not None:
            out.append(m)
    return out


def organize(input_folder: str, output_dir: str, write_xlsx: bool = True) -> dict:
    """Copia arquivos para output_dir/{bounty_ko,mystery,vanilla}. Retorna contagens."""
    target_dirs = {
        "Bounty/KO": os.path.join(output_dir, "bounty_ko"),
        "Mystery":   os.path.join(output_dir, "mystery"),
        "Vanilla":   os.path.join(output_dir, "vanilla"),
    }
    for d in target_dirs.values():
        os.makedirs(d, exist_ok=True)

    metadata = collect_metadata(input_folder)
    tournaments: dict[str, dict] = defaultdict(
        lambda: {"files": [], "tname": "", "buyin": "", "totalbuyin": "",
                 "startdate": "", "tablesize": ""})
    for m in metadata:
        key = m["tcode"] or m["file"]
        t = tournaments[key]
        t["files"].append(m)
        for k in ("tname", "buyin", "totalbuyin", "startdate", "tablesize"):
            if not t[k]:
                t[k] = m[k]

    rows = []
    counts = {"Bounty/KO": 0, "Mystery": 0, "Vanilla": 0}
    file_counts = {"Bounty/KO": 0, "Mystery": 0, "Vanilla": 0}
    for tcode, t in tournaments.items():
        tipo = classify(t["buyin"], t["tname"])
        counts[tipo] += 1
        file_counts[tipo] += len(t["files"])
        for fm in t["files"]:
            shutil.copy2(fm["path"], os.path.join(target_dirs[tipo], fm["file"]))
        rows.append({
            "tcode": tcode, "tname": t["tname"], "buyin": t["buyin"],
            "totalbuyin": t["totalbuyin"], "startdate": t["startdate"],
            "tablesize": t["tablesize"], "tipo": tipo, "n_files": len(t["files"]),
        })

    if write_xlsx:
        try:
            _write_xlsx(rows, os.path.join(output_dir, "torneios_classificados.xlsx"))
        except ImportError:
            print("openpyxl nao instalado — pulando geracao do xlsx")

    return {"tournaments": counts, "files": file_counts, "rows": rows}


def _write_xlsx(rows: list[dict], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    order = {"Bounty/KO": 0, "Mystery": 1, "Vanilla": 2}
    rows = sorted(rows, key=lambda r: (order[r["tipo"]], r["startdate"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Torneios"
    headers = ["Tipo", "Tournament Code", "Nome do Torneio", "Buyin",
               "Total Buyin", "Data Inicial", "Mesa", "Qtd. Arquivos"]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    fills = {
        "Bounty/KO": PatternFill("solid", start_color="FCE4D6"),
        "Mystery":   PatternFill("solid", start_color="FFF2CC"),
        "Vanilla":   PatternFill("solid", start_color="E2EFDA"),
    }
    arial = Font(name="Arial", size=11)

    for r in rows:
        ws.append([r["tipo"], r["tcode"], r["tname"], r["buyin"],
                   r["totalbuyin"], r["startdate"], r["tablesize"], r["n_files"]])
        idx = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col)
            cell.fill = fills[r["tipo"]]
            cell.border = border
            cell.font = arial

    widths = [12, 17, 42, 28, 14, 22, 7, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="Pasta com XMLs originais")
    ap.add_argument("--output-dir", default=".",
                    help="Onde criar pastas bounty_ko/mystery/vanilla")
    ap.add_argument("--no-xlsx", action="store_true")
    args = ap.parse_args()
    result = organize(args.input, args.output_dir, write_xlsx=not args.no_xlsx)
    print("Torneios:", result["tournaments"])
    print("Arquivos:", result["files"])


if __name__ == "__main__":
    main()
